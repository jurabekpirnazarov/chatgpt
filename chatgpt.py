import os
import openai
import pandas as pd
import datetime
from openpyxl import Workbook, load_workbook
import time

openai.api_key = 'your key'

for i in range(2900, 3847):
  df = pd.read_excel('/content/drive/MyDrive/All questions.xlsx')
  time.sleep(10)
  try: 
    print(i)
    #chatgpt api 
    completion1 = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
        {"role": "user", "content":  f'{df.iloc[i].Questions} make 6 answers to the question more than 3 sentences'}]
        )
    current_time = datetime.datetime.now()
    #to save excell sheets
    workbook = load_workbook(filename='/content/drive/MyDrive/Questions to the answer.xlsx')
    worksheet = workbook.active
    cell = worksheet.cell(row=i+2, column=4)
    cell.value = completion1.choices[0].message.content
    workbook.save(filename='/content/drive/MyDrive/Questions to the answer.xlsx')
    print("Current time: " + current_time.strftime("%H:%M:%S")+'==> done')
  except:
    time.sleep(21)
    print(i)
    completion1 = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
        {"role": "user", "content":  f'{df.iloc[i].Questions} make 6 answers to the question more than 3 sentences'}]
        )
    #to save excell sheets
    current_time = datetime.datetime.now()
    workbook = load_workbook(filename='/content/drive/MyDrive/Questions to the answer.xlsx')
    worksheet = workbook.active
    cell = worksheet.cell(row=i+2, column=4)
    cell.value = completion1.choices[0].message.content
    workbook.save(filename='/content/drive/MyDrive/Questions to the answer.xlsx')
    print("Current time: " + current_time.strftime("%H:%M:%S")+'==> done2')
    continue

  print(f'{i} Tugadi')
    # waits for 5 seconds---600. 2088.2181

  if i %100 == 0:

    current_time = datetime.datetime.now()
    print("Current time: " + current_time.strftime("%H:%M:%S"))
  # after each 100 units, the data collects to the google drive disk 1936
print('==============FINISHED!!!!!!!!!!!!!!!!!!==============')
